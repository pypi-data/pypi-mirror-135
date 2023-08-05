"""
Created on 18.05.2021

@author: baier
"""
import datetime
import locale
import openpyxl as px
import pandas as pd
from openpyxl import styles

now = (datetime.datetime.now()).strftime('%d.%m.%Y-%H:%M:%S')
locale.setlocale(locale.LC_ALL, 'de_DE')

greyFill = styles.PatternFill(start_color='00808080',
                              end_color='00808080',
                              fill_type='solid')


def formatExcelcshsm(path: str):
    # ## PANDAS ###

    df = pd.read_csv(path, header=1, on_bad_lines="skip")  # load csv and set header == line 2
    df = df.drop(df.index[len(df) - 1])  # drop last row

    # rearrange columns
    cols = ['client_ref_num', 'security', 'isin', 'sedol', 'cusip', 'quantity', 'price', 'currency_code', 'comm_fee',
            'principal',
            'signed_total_curr', 'trade_date', 'settle_date', 'broker_name', 'company_name', 'common_account_name',
            'common_account_id', 'input_date', 'identifier', 'asset_type',
            'tkt_state', 'txn_type', 'cash_balance', 'base_ccy', 'base_fx', 'base_signed_total_curr', 'tax_fee', 'exch_fee',
            'sec_fee',
            'int_acrued', 'other_fees', 'option_reg_fee', 'gross_amt', 'withholding', 'bb_ticker', 'occ_symbol', 'house_id']
    df = df[cols]

    # delete opening/closing positions AND sort by security
    df = df[df['security'] != 'Opening Balance']
    df = df[df['security'] != 'Closing Balance']
    # df = df.sort_index()
    df = df.sort_values(by='security')

    # format columns into length of col name
    new_file = path[:-4] + '_edited.xlsx'
    writer = pd.ExcelWriter(new_file)  # options={'strings_to_numbers': True}
    df.to_excel(writer, sheet_name='Report', index=False, na_rep='NaN')
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column)) + 2
        col_idx = df.columns.get_loc(column)
        writer.sheets['Report'].set_column(col_idx, col_idx, column_width)

    writer.save()

    # ## OPENPYXL ###
    wb = px.load_workbook(new_file)
    ws = wb.active

    cols = ws.max_column
    ws.auto_filter.ref = ws.dimensions  # filter for excel

    for cell in ws['F']: cell.number_format = '#,##'  # format cells to thousands point and integer
    for cell in ws['J']: cell.number_format = '#,##'  # format cells to thousands point and integer
    for cell in ws['K']: cell.number_format = '#,##'  # format cells to thousands point and integer
    # ws.cell(1,1).value = 'client_ref_num'           # rename columns
    # ws.cell(1,13).value = 'broker_name'             # rename columns

    for col in range(1, cols + 1):
        ws.cell(1, col).fill = greyFill

    wb.save(new_file)

    return print('At', now + ':', path[-9:], 'successfully edited')


def formatExcelposnd(path: str):
    # ## PANDAS ###

    df = pd.read_csv(path, header=1, on_bad_lines="skip")  # load csv and set header == line 2
    df = df.drop(df.index[len(df) - 1])  # drop last row

    # rearrange columns
    cols = ['security', 'isin', 'sedol', 'cusip', 'market', 'td_quantity', 'sd_quantity', 'live_price', 'asset_type',
            'identifier', 'company_name', 'common_account_name', 'common_account_id', 'acct_type', 'position_date',
            'id_quantity', 'currency_code',
            'px_multiplier', 'td_mv', 'sd_mv', 'id_mv', 'base_ccy', 'base_fx', 'td_mv_base', 'sd_mv_base', 'id_mv_base',
            'ticker', 'option_key', 'house_id']
    df = df[cols]

    # change ***WTS, WTS and CORP to Warrents
    df.loc[df.security.str[:3] == '***', 'security'] = df.security.str[3:]
    df.loc[df.asset_type == 'CORP', 'asset_type'] = 'WARRANT'
    df.loc[df.security.str[:3] == 'WTS', 'asset_type'] = 'WARRANT'

    # df = df.sort_index()
    df = df.sort_values(['asset_type', 'security'])

    # double reset_index to get only one index column
    # df = df.reset_index(drop=True)
    # df = df.reset_index()

    # format columns into length of col name
    new_file = path[:-4] + '_edited.xlsx'
    writer = pd.ExcelWriter(new_file)  # options={'strings_to_numbers': True}
    df.to_excel(writer, sheet_name='Report', index=False, na_rep='NaN')
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column)) + 2
        col_idx = df.columns.get_loc(column)
        writer.sheets['Report'].set_column(col_idx, col_idx, column_width)

    writer.save()

    # ## OPENPYXL ###
    wb = px.load_workbook(new_file)
    ws = wb.active

    cols = ws.max_column
    ws.auto_filter.ref = ws.dimensions  # filter for excel

    for cell in ws['F']: cell.number_format = '#,##'  # format cells to thousands point and integer
    for cell in ws['G']: cell.number_format = '#,##'  # format cells to thousands point and integer

    for col in range(1, cols + 1):
        ws.cell(1, col).fill = greyFill

    ws.freeze_panes = "A2"

    wb.save(new_file)

    return print('At', now + ':', path[-9:], 'successfully edited')

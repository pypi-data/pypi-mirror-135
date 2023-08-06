#!/usr/bin/python3

# DIDIS - Desy ITk Database Interaction Script -- DESY's very own framework for interacting with the ITk Production Database
# Based on itkdb: https://gitlab.cern.ch/atlas-itk/sw/db/itkdb
# Created: 2021/11/17, Updated: 2021/11/19
# Written by Maximilian Felix Caspar, DESY HH


from loguru import logger
import argh
import yaml
from openpyxl import load_workbook
from copy import deepcopy
from didis.didis import lookup


def excel(excelFile: "Excel file containing the test data",
          jobYAML: "YAML containing the job definition"):
    logger.info(f"Executing job {jobYAML} on {excelFile}")
    job = yaml.load(open(jobYAML, 'r'))
    print(job)
    workbook = load_workbook(filename=excelFile)
    sheet = workbook[job['sheet']]
    # Get aceptance regions from excel file
    acceptance = {k:
                  {
                      "min": float(sheet.cell(job["reference"]["nominal"], job["frame"]["data"][k]).value)
                      + float(sheet.cell(job["reference"]["lower"], job["frame"]["data"][k]).value),

                      "max": float(sheet.cell(job["reference"]["nominal"], job["frame"]["data"][k]).value)
                      + float(sheet.cell(job["reference"]["upper"], job["frame"]["data"][k]).value)
                  }
                  for k in job["frame"]["data"] if k != job["lookupKey"]}
    print(acceptance)
    for i in range(10000):
        line = job["frame"]["offset"] + i
        df = deepcopy(job["frame"]["data"])
        print(df)
        for k in df:
            df[k] = sheet.cell(line, df[k]).value
        if df[job["lookupKey"]] is None:
            break
        component = lookup(project=job["project"], lookupKey=job["lookupKey"],
                           value=df[job["lookupKey"]], componentType=job["componentType"])
        if i > 4:
            break


def main():
    parser = argh.ArghParser()
    parser.add_commands([excel])
    parser.dispatch()


if __name__ == '__main__':
    main()

from openpyxl import Workbook
from scrapy.xcc_items.factoryitems import FactoryMaterialItem, FactoryBaikeItem
'''
author: tieyongjie
'''

class ExcelPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.wb.create_sheet("ware_detail")
        self.ws = self.wb["ware_detail"]
        self.ws.append(
            ['sources', 'url', 'title', 'category_name', 'category_id', 'list_json', 'brand_id', 'create_time',
             'creator', 'brand_name', 'img_url', 'pdf_url'])
        self.wb.create_sheet("wiki")
        self.ws_2 = self.wb["wiki"]
        self.ws_2.append(
            ['sources', 'url', 'title', 'brand_name', 'brand_id', 'category_name', 'category_id', 'create_time',
             'creator', 'application', 'feature', 'standard', 'overview','description'])

    def process_item(self, item, spider):
        if isinstance(item, FactoryMaterialItem):
            self.ws.append([item['sources'], item['url'], item['title'], item['category_name'],
                            item['category_id'], item['list_json'], item['brand_id'],
                            item['create_time'], item['creator'], item['brand_name'], item['img_url'], item['pdf_url']])
        if isinstance(item, FactoryBaikeItem):
            self.ws_2.append([item['sources'], item['url'], item['title'], item['brand_name'],
                              item['brand_id'], item['category_name'], item['category_id'],
                              item['create_time'], item['creator'], item['application'], item['feature'],
                              item['standard'],item['overview'],item['description']])
        return item

    def __del__(self):
        # 调用__del__() 销毁对象，释放其空间
        self.wb.save('wuliao.xlsx')
